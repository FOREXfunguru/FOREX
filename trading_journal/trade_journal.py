import pandas as pd
import numpy as np
import logging
import math
import re
from typing import Tuple, List

import openpyxl
from openpyxl import Workbook

from trading_journal.open_trade import UnawareTrade
from params import tjournal_params


tj_logger = logging.getLogger(__name__)
tj_logger.setLevel(logging.INFO)


class TradeJournal(object):
    """
    Constructor

    Class variables:
        url: path to the .xlsx file with the trade journal
        worksheet: Name of the worksheet that will be used to create 
                   the object. i.e. 'trading_journal'
    """
    __slots__ = ['url', 'worksheet', 'df']

    def __init__(self, url: str, worksheet: str):
        self.url = url
        self.worksheet = worksheet

        # read-in the 'trading_journal' worksheet from a .xlsx file into a
        # pandas dataframe
        try:
            xls_file = pd.ExcelFile(url)
            df = xls_file.parse(worksheet, converters={
                "id": str,
                "start": str,
                "end": str,
                "trend_i": str})
            df = df.dropna(how='all')

            if df.empty is True:
                raise Exception(f"No trades fetched for url:{self.url} and "
                                "worksheet:{self.worksheet}")
            # replace n.a. string by NaN
            df = df.replace("n.a.", np.NaN)
            # remove trailing whitespaces from col names
            df.columns = df.columns.str.rstrip()
            self.df = df
        except FileNotFoundError:
            wb = Workbook()
            wb.create_sheet(worksheet)
            wb.save(str(self.url))

    def fetch_trades(self, init_clist: bool=False) -> List[UnawareTrade]:
        """Function to fetch a list of Trade objects.
        
        Args:
            init_clist: If True, then clist and clist_tm will be initialised.
        """

        trade_list, args = [], {}
        for _, row in self.df.iterrows():
            if isinstance(row['id'], float):
                continue
            elms = re.split(r'\.| ', row['id'])
            assert len(elms) >= 1, "Error parsing the trade id"
            pair = elms[0]
            args = {'pair': pair, **row}
            t = UnawareTrade(**args, init_clist=init_clist)
            trade_list.append(t)

        return trade_list

    def win_rate(self, strats: str) -> Tuple[int, int, float]:
        '''Calculate win rate and pips balance
        for this TradeJournal. If outcome attrb is not
        defined then it will invoke the run_trade method
        on each particular trade

        Arguments:
            strats : Comma-separated list of strategies to analyse:
                     i.e. counter,counter_b1

        Returns:
            number of successes
            number of failures
            balance of pips in this TradeList
        '''

        strat_l = strats.split(",")
        number_s = number_f = tot_pips = 0
        for _, row in self.df.iterrows():
            pair = row['id'].split(" ")[0]
            args = {'pair': pair, **row}
            t = UnawareTrade(**args, init_clist=True)
            if t.strat not in strat_l:
                continue
            if not hasattr(t, 'outcome') or math.isnan(t.outcome):
                t.initialise(expires=1)
                t.run()
            if t.outcome == 'success':
                number_s += 1
            elif t.outcome == 'failure':
                number_f += 1
            tot_pips += t.pips
        tot_pips = round(tot_pips, 2)
        tot_trades = number_s+number_f
        perc_wins = round(number_s*100/tot_trades, 2)
        perc_losses = round(number_f*100/tot_trades, 2)
        print("Tot number of trades: {0}\n-------------".format(tot_trades))
        print("Win trades: {0}; Loss trades: {1}".format(number_s, number_f))
        print("% win trades: {0}; % loss trades: {1}".format(perc_wins,
                                                             perc_losses))
        print("Pips balance: {0}".format(tot_pips))

        return number_s, number_f, tot_pips

    def write_tradelist(self, trade_list: List[UnawareTrade],
                        sheet_name: str) -> None:
        """Write the TradeList to the Excel spreadsheet
        pointed by the trade_journal.

        Arguments:
            trade_list : List of Trade objects
            sheet_name : worksheet name"""
        colnames = tjournal_params.colnames.split(",")
        data = []
        for t in trade_list:
            id = f"{t.pair} {t.start.strftime('%Y%m%d')}"
            t.id = id
            row = []
            for key in colnames:
                # some keys are not defined for some of the Trade
                # objects
                if hasattr(t, key):
                    if key in ["SL", "entry", "TP"]:
                        area = getattr(t, key)
                        row.append(area.price)
                    else:
                        row.append(getattr(t, key))
                else:
                    row.append("n.a.")
            data.append(row)
                
        df = pd.DataFrame(data, columns=colnames)

        writer = pd.ExcelWriter(self.url, engine='openpyxl', mode='a',
                                if_sheet_exists='replace')
        writer.workbook = openpyxl.load_workbook(self.url)
        tj_logger.info("Creating new worksheet with trades with name: {0}".
                       format(sheet_name))
        df.to_excel(writer, sheet_name)
        writer.save()
