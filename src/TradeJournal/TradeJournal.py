import pandas as pd
import pdb
import re
import datetime
import warnings
from TradeJournal.Trade import Trade
from openpyxl import load_workbook
from Pattern.Counter import Counter


class TradeJournal(object):
    '''
    Constructor

    Class variables
    ---------------
    
    url: path to the .xlsx file with the trade journal
    worksheet: str, Required
               Name of the worksheet that will be used to create the object. i.e. trading_journal
    '''

    def __init__(self, url, worksheet):
        self.url=url
        self.worksheet=worksheet
        #read-in the 'trading_journal' worksheet from a .xlsx file into a pandas dataframe
        xls_file = pd.ExcelFile(url)
        df = xls_file.parse(worksheet,converters={'start': str, 'end': str, 'trend_i': str})
        self.df=df

    def fetch_trades(self,strat=None):
        '''
        This function will fetch the trades that are in this TradingJournal and will create an independent
        Trade object for each record

        Parameters
        ----------
        strat : String, Optional
                If defined, then select all trades of a certain type.
                Possible values are: 'counter'

        Returns
        ------
        Nothing

        '''

        trade_list=[]
        for index,row in self.df.iterrows():
            if strat is not None:
                if strat!=row['strat']:
                    continue
            #get pair from id
            pair=row['id'].split(' ')[0]

            attrbs={}
            for items in row.iteritems():
                attrbs[items[0]]=items[1]

            c =Counter(pair=pair,**attrbs)

            c.init_feats()

            p = re.compile('clist_')
            attrbs1={}

            for attr, value in c.__dict__.items():
                if p.match(attr):continue
                if type(value) is list:
                    attrbs1[attr]=len(value)
                else:
                    attrbs1[attr] = value

            t=Trade(id=row['id'], strat=row['strat'],**attrbs1)
            trade_list.append(t)

        return trade_list

    def write_trades(self, trade_list, colnames):
        '''
        Write the trade_list to the Excel spreadsheet
        pointed by the TradeJournal

        Parameters
        ----------
        trade_list : list, Required
                     List with Trade objects
        colnames : list, Required
                    Column names that will control the order
                    of the columns

        Returns
        -------
        Nothing
        '''

        data=[]
        for t in trade_list:
            row=[]
            for a in colnames:
                row.append(getattr(t, a))
            data.append(row)

        df = pd.DataFrame(data, columns=col_names)

        book = load_workbook(self.url)
        writer = pd.ExcelWriter(self.url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, 'calculated_trades')
        writer.save()

    def add_trend_momentum(self):
        '''
        This function will add a new worksheet named 'trend_momentum' to the .xlsx file
        For this, the function will perform some queries to the Oanda's REST API and will
        parse the results
        '''

        trade_list=self.fetch_trades()
        data=[]
        for trade in trade_list:
            print("Processing: {0}".format(trade.start.strftime("%d-%m-%y")))
            if trade.timeframe=="H10":
                warnings.warn("Timeframe format: {0} is not valid. Skipping".format(trade.timeframe))
                continue
            if trade.timeframe=="na":
                warnings.warn("Timeframe is not defined. Skipping")
                continue
            row=[]
            row.extend((trade.start, trade.end, trade.pair, trade.type, trade.timeframe, trade.outcome))
            cl=trade.fetch_candlelist()
            if cl.len==1:
                warnings.warn("CandleList has only one candle. Skipping")
                continue
            cl.calc_binary_seq(merge=True)
            cl.calc_number_of_0s(norm=True)
            cl.calc_number_of_doubles0s(norm=True)
            cl.calc_longest_stretch()
            cl.get_entropy(norm=True)

            for i in ['high','low','open','close','colour','merge']:
                row.append("'{0}'".format(cl.seq[i]))
            for i in ['high','low','open','close','colour','merge']:
                row.append(cl.number_of_0s[i])
            for i in ['high','low','open','close','colour']:
                row.append(cl.longest_stretch[i])
            for i in ['high','low','open','close','colour']:
                row.append(cl.entropy[i])
            row.extend((cl.highlow_double0s,cl.openclose_double0s))
            data.append(row)

        columns=['start','end','pair','type','timeframe','outcome','high','low','open','close','colour','merge',
                 'No_of_0s_high','No_of_0s_low','No_of_0s_open','No_of_0s_close','No_of_0s_colour','No_of_0s_merge',
                 'stretch_high','stretch_low','stretch_open','stretch_close','stretch_colour',
                 'ent_high','ent_low','ent_open','ent_close','ent_colour',
                 'No_of_double0_higlow','No_of_double0_openclose']
        df=pd.DataFrame(data,columns=columns)
        book = load_workbook(self.url)
        writer = pd.ExcelWriter(self.url, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer,'trend_momentum')
        writer.save()
